import os
import base64
import tempfile
import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from fpdf import FPDF
from statsmodels.tsa.holtwinters import ExponentialSmoothing
from statsmodels.tsa.seasonal import seasonal_decompose
 
st.set_page_config(page_title="Financial Analytics AI", layout="wide")
 
try:
    GEMINI_API_KEY = st.secrets["GEMINI_API_KEY"]
except (KeyError, FileNotFoundError):
    GEMINI_API_KEY = ""

def detect_anomalies(series: pd.Series, threshold: float = 2.0) -> pd.Series:
    rolling_mean = series.rolling(window=3, min_periods=1).mean()
    rolling_std  = series.rolling(window=3, min_periods=1).std().fillna(1)
    return ((series - rolling_mean) / rolling_std).abs() > threshold

def get_ai_commentary(kpis: dict) -> tuple:
    """Call Gemini API for a Polish executive summary.
    Returns (text, error_message) — one will always be empty."""
    if not GEMINI_API_KEY:
        return "", "Brak klucza API. Dodaj GEMINI_API_KEY do .streamlit/secrets.toml."
    try:
        import google.generativeai as genai
        genai.configure(api_key=GEMINI_API_KEY)
        model = genai.GenerativeModel("gemini-2.0-flash")
        prompt = (
            "Jesteś zwięzłym asystentem CFO. Na podstawie poniższych miesięcznych "
            "KPI napisz dokładnie 3 krótkie, wnikliwe zdania po polsku jako podsumowanie "
            "dla zarządu. Odnoś się konkretnie do liczb. Nie używaj punktorów.\n\n"
            f"KPI: {kpis}"
        )
        response = model.generate_content(prompt)
        return response.text.strip(), ""
    except Exception as e:
        return "", f"Błąd API: {e}"
 
def create_pdf(df, pred, runway, fig_mc, ai_text=""):
    pdf = FPDF()
    pdf.add_page()
    import plotly.io as pio
 
    pdf.set_font("Arial", "B", 16)
    pdf.cell(190, 10, "Raport Finansowy - Analiza Predykcyjna", ln=True, align="C")
    pdf.ln(8)
 
    if ai_text:
        pdf.set_font("Arial", "I", 11)
        pdf.multi_cell(190, 7, ai_text)
        pdf.ln(4)
 
    pdf.set_font("Arial", "", 12)
    pdf.cell(190, 10, f"Prognoza przychodow (t+1): {pred.iloc[0]:,.2f} PLN", ln=True)
    pdf.cell(190, 10, f"Runway (bezpieczenstwo): {runway} miesiecy", ln=True)
    pdf.ln(4)
 
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
        tmp_path = tmp.name
    try:
        pio.write_image(fig_mc, tmp_path, engine="kaleido")
        pdf.image(tmp_path, x=10, y=None, w=180)
    finally:
        os.unlink(tmp_path)
    pdf.ln(4)
 
    pdf.set_font("Arial", "B", 10)
    for col in ["Data", "Przychody", "Wydatki"]:
        pdf.cell(60, 10, col, border=1)
    pdf.ln()
    pdf.set_font("Arial", "", 10)
    for i in range(1, 6):
        pdf.cell(60, 10, str(df.index[-i].date()), border=1)
        pdf.cell(60, 10, f"{df['Przychody'].iloc[-i]:,.0f}", border=1)
        pdf.cell(60, 10, f"{df['Wydatki'].iloc[-i]:,.0f}", border=1, ln=True)
 
    return pdf.output(dest="S").encode("latin-1", "replace")
 
def to_excel(df: pd.DataFrame) -> bytes:
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
 
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dane finansowe"
 
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True)
 
    df_export = df.copy().reset_index()
    df_export["Data"] = df_export["Data"].dt.strftime("%Y-%m-%d")
    df_export["Marza (%)"] = (
        (df_export["Przychody"] - df_export["Wydatki"]) / df_export["Przychody"] * 100
    ).round(1)
 
    for r_idx, row in enumerate(dataframe_to_rows(df_export, index=False, header=True), 1):
        ws.append(row)
        for cell in ws[r_idx]:
            cell.alignment = Alignment(horizontal="center")
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
 
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4
 
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
 
st.title("Financial Intelligence Dashboard")
 
uploaded_file = st.sidebar.file_uploader("Wgraj plik CSV", type="csv")
 
st.sidebar.divider()
st.sidebar.subheader("Filtr okresu")
date_filter = st.sidebar.selectbox(
    "Pokaż dane z ostatnich:",
    ["Wszystkie", "6 miesięcy", "12 miesięcy", "24 miesiące"],
)
 
st.sidebar.divider()
st.sidebar.subheader("Symulator What-If")
revenue_shock = st.sidebar.slider("Zmiana przychodów (%)", -50, 50, 0, 5)
burn_shock    = st.sidebar.slider("Zmiana kosztów (%)",    -50, 50, 0, 5)
 
st.sidebar.divider()
st.sidebar.subheader("Komentarz AI")

generate_ai_btn = st.sidebar.button("Generuj / Odśwież komentarz AI", disabled=not bool(GEMINI_API_KEY))

if "ai_comment" not in st.session_state:
    st.session_state.ai_comment = ""
if "ai_error" not in st.session_state:
    st.session_state.ai_error = ""

if uploaded_file:
    try:
        df_raw = pd.read_csv(uploaded_file)
        df_raw["Data"] = pd.to_datetime(df_raw["Data"])
        df_raw = df_raw.sort_values("Data").set_index("Data")
 
        months_map = {"6 miesięcy": 6, "12 miesięcy": 12, "24 miesiące": 24}
        if date_filter in months_map:
            cutoff = df_raw.index[-1] - pd.DateOffset(months=months_map[date_filter])
            df = df_raw[df_raw.index >= cutoff].copy()
        else:
            df = df_raw.copy()
 
        sp     = min(4, len(df) // 2)
        model  = ExponentialSmoothing(df["Przychody"], seasonal="add", seasonal_periods=sp).fit()
        pred   = model.forecast(6)
        pred_adj = pred * (1 + revenue_shock / 100)
 
        rev_anomalies = detect_anomalies(df["Przychody"])
        exp_anomalies = detect_anomalies(df["Wydatki"])
        any_anomaly   = rev_anomalies | exp_anomalies
        n_anomalies   = int(any_anomaly.sum())
 
        last_rev  = df["Przychody"].iloc[-1]
        prev_rev  = df["Przychody"].iloc[-2]
        rev_diff  = ((last_rev - prev_rev) / prev_rev) * 100
 
        burn_base = df["Wydatki"].mean()
        burn_adj  = burn_base * (1 + burn_shock / 100)
        runway    = round(
            (df["Przychody"].sum() * (1 + revenue_shock / 100)
             - df["Wydatki"].sum() * (1 + burn_shock / 100)) / burn_adj, 1
        )
 
        df["Zysk"]      = df["Przychody"] - df["Wydatki"]
        df["Marza %"]   = (df["Zysk"] / df["Przychody"] * 100).round(1)
        df["MoM Rev %"] = df["Przychody"].pct_change().mul(100).round(1)
        df["MoM Exp %"] = df["Wydatki"].pct_change().mul(100).round(1)
 
        kpis = {
            "ostatni_przychod_PLN": round(last_rev),
            "zmiana_MoM_%": round(rev_diff, 1),
            "sredni_burn_rate_PLN": round(burn_adj),
            "runway_msc": runway,
            "anomalie": n_anomalies,
            "prognoza_t1_PLN": round(pred_adj.iloc[0]),
            "srednia_marza_%": round(df["Marza %"].mean(), 1),
        }

        if generate_ai_btn:
            with st.spinner("Generowanie komentarza AI..."):
                ai_text, ai_error = get_ai_commentary(kpis)
                st.session_state.ai_comment = ai_text
                st.session_state.ai_error = ai_error

        if st.session_state.ai_error:
            st.warning(f"Komentarz AI niedostępny: {st.session_state.ai_error}")
        elif st.session_state.ai_comment:
            st.info(f"**Komentarz AI:** {st.session_state.ai_comment}")
 
        st.subheader("Kluczowe Wskazniki")
        k1, k2, k3, k4, k5 = st.columns(5)
        k1.metric("Ostatni Przychod",  f"{last_rev:,.0f} PLN", f"{rev_diff:.1f}%")
        k2.metric("Sr. Burn Rate",     f"{burn_adj:,.0f} PLN", f"{burn_shock:+d}%" if burn_shock else None)
        k3.metric("Runway (msc)",      f"{runway}")
        k4.metric("Sr. Marza",         f"{df['Marza %'].mean():.1f}%")
        k5.metric("Anomalie",          f"{n_anomalies}",
                  delta="sprawdz!" if n_anomalies > 0 else "brak",
                  delta_color="inverse" if n_anomalies > 0 else "normal")
 
        if n_anomalies > 0:
            dates_str = df.index[any_anomaly].strftime("%Y-%m").tolist()
            st.warning(f"Anomalie wykryte w: {', '.join(dates_str)}")
 
        st.divider()

        r1c1, r1c2 = st.columns(2)

        with r1c1:
            st.subheader("Przychody vs Wydatki")
            fig_rev = go.Figure()
            fig_rev.add_trace(go.Scatter(
                x=df.index, y=df["Przychody"], name="Przychody",
                line=dict(color="rgba(0,200,100,0.9)", width=2),
            ))
            fig_rev.add_trace(go.Scatter(
                x=df.index, y=df["Wydatki"], name="Wydatki",
                fill="tonexty", fillcolor="rgba(0,200,100,0.1)",
                line=dict(color="rgba(255,80,80,0.9)", width=2),
            ))
            anom_idx = df.index[any_anomaly]
            if len(anom_idx):
                fig_rev.add_trace(go.Scatter(
                    x=anom_idx, y=df.loc[anom_idx, "Przychody"],
                    mode="markers", name="Anomalia",
                    marker=dict(color="red", size=10, symbol="x"),
                ))
            fig_rev.update_layout(template="plotly_dark", height=350,
                                   legend=dict(orientation="h", y=1.1))
            st.plotly_chart(fig_rev, use_container_width=True)
 
        with r1c2:
            st.subheader("Monte Carlo + Przedzial Ufnosci")
            n_sim  = 200
            sigma  = df["Przychody"].pct_change().std()
            dates  = pd.date_range(df.index[-1], periods=7, freq="ME")[1:]
 
            sim_matrix = np.array([
                pred_adj.values * (1 + np.random.normal(0, sigma, 6))
                for _ in range(n_sim)
            ])
            p10 = np.percentile(sim_matrix, 10, axis=0)
            p90 = np.percentile(sim_matrix, 90, axis=0)
 
            fig_mc = go.Figure()
            for path in sim_matrix[:40]:
                fig_mc.add_trace(go.Scatter(
                    x=dates, y=path, mode="lines",
                    line=dict(width=0.4, color="rgba(100,100,255,0.15)"),
                    showlegend=False,
                ))
            fig_mc.add_trace(go.Scatter(
                x=list(dates) + list(dates[::-1]),
                y=list(p90) + list(p10[::-1]),
                fill="toself", fillcolor="rgba(255,165,0,0.15)",
                line=dict(color="rgba(255,165,0,0)"),
                name="Przedzial 10-90%",
            ))
            fig_mc.add_trace(go.Scatter(
                x=dates, y=pred_adj, name="Prognoza",
                line=dict(color="orange", width=3),
            ))
            fig_mc.update_layout(template="plotly_dark", height=350,
                                  legend=dict(orientation="h", y=1.1))
            st.plotly_chart(fig_mc, use_container_width=True)
 
        st.divider()
 
        r2c1, r2c2 = st.columns(2)
 
        with r2c1:
            st.subheader("Marza Zysku (%)")
            colors = [
                "rgba(0,200,100,0.8)" if v >= 0 else "rgba(255,80,80,0.8)"
                for v in df["Marza %"]
            ]
            fig_margin = go.Figure(go.Bar(
                x=df.index, y=df["Marza %"],
                marker_color=colors,
                text=df["Marza %"].apply(lambda x: f"{x:.1f}%"),
                textposition="outside",
            ))
            fig_margin.update_layout(template="plotly_dark", height=320,
                                      yaxis_title="%", showlegend=False)
            st.plotly_chart(fig_margin, use_container_width=True)
 
        with r2c2:
            st.subheader("Waterfall — Skumulowane Cash-Flow")
            monthly = df["Zysk"].values
            fig_wf = go.Figure(go.Waterfall(
                x=df.index.strftime("%Y-%m"),
                y=monthly,
                measure=["relative"] * len(monthly),
                connector=dict(line=dict(color="rgba(255,255,255,0.2)")),
                increasing=dict(marker_color="rgba(0,200,100,0.8)"),
                decreasing=dict(marker_color="rgba(255,80,80,0.8)"),
            ))
            fig_wf.update_layout(template="plotly_dark", height=320, yaxis_title="PLN")
            st.plotly_chart(fig_wf, use_container_width=True)
 
        st.divider()
 
        r3c1, r3c2 = st.columns([2, 1])
 
        with r3c1:
            st.subheader("Dekompozycja Sezonowa (Przychody)")
            try:
                decomp = seasonal_decompose(df["Przychody"], model="additive", period=sp)
                fig_decomp = go.Figure()
                for name, series, color in [
                    ("Trend",      decomp.trend,    "orange"),
                    ("Sezonowosc", decomp.seasonal, "cyan"),
                    ("Rezydua",    decomp.resid,    "rgba(180,180,180,0.6)"),
                ]:
                    fig_decomp.add_trace(go.Scatter(
                        x=df.index, y=series, name=name,
                        line=dict(color=color, width=2),
                    ))
                fig_decomp.update_layout(template="plotly_dark", height=320,
                                          legend=dict(orientation="h", y=1.1))
                st.plotly_chart(fig_decomp, use_container_width=True)
            except Exception:
                st.info("Za malo danych do dekompozycji sezonowej (wymagane min. 2 okresy).")
 
        with r3c2:
            st.subheader("Struktura Przychodu")
            last_exp = df["Wydatki"].iloc[-1] * (1 + burn_shock / 100)
            profit   = max(last_rev * (1 + revenue_shock / 100) - last_exp, 0)
            fig_pie  = px.pie(
                names=["Zysk", "Koszty"], values=[profit, last_exp], hole=0.4,
                color_discrete_map={"Zysk": "rgba(0,200,100,0.8)", "Koszty": "rgba(255,80,80,0.8)"},
            )
            fig_pie.update_layout(template="plotly_dark", height=320)
            st.plotly_chart(fig_pie, use_container_width=True)
 
        st.divider()
 
        r4c1, r4c2 = st.columns(2)
 
        with r4c1:
            st.subheader("Zmiana Miesieczna (MoM)")
            mom_df = df[["Przychody", "Wydatki", "Zysk", "MoM Rev %", "MoM Exp %", "Marza %"]].copy()
            mom_df.index = mom_df.index.strftime("%Y-%m")
 
            def color_signed(val):
                if pd.isna(val):
                    return ""
                return "color: rgba(0,220,100,0.9)" if val >= 0 else "color: rgba(255,80,80,0.9)"
 
            styled = (
                mom_df.style
                .format({
                    "Przychody":  "{:,.0f}",
                    "Wydatki":    "{:,.0f}",
                    "Zysk":       "{:,.0f}",
                    "MoM Rev %":  "{:+.1f}%",
                    "MoM Exp %":  "{:+.1f}%",
                    "Marza %":    "{:.1f}%",
                })
                .map(color_signed, subset=["MoM Rev %", "MoM Exp %", "Marza %"])
            )
            st.dataframe(styled, use_container_width=True, height=300)
 
        with r4c2:
            st.subheader("Porownanie scenariuszy")
            runway_base = round(
                (df["Przychody"].sum() - df["Wydatki"].sum()) / burn_base, 1
            )
            last_rev_adj  = last_rev * (1 + revenue_shock / 100)
            last_margin_adj = (
                (last_rev_adj - last_exp) / last_rev_adj * 100
            ) if last_rev_adj else 0
 
            compare = pd.DataFrame({
                "Scenariusz":         ["Bazowy", "What-If"],
                "Burn Rate (PLN)":    [f"{burn_base:,.0f}", f"{burn_adj:,.0f}"],
                "Runway (msc)":       [runway_base, runway],
                "Prognoza t+1 (PLN)": [f"{pred.iloc[0]:,.0f}", f"{pred_adj.iloc[0]:,.0f}"],
                "Marza (last, %)":    [
                    f"{df['Marza %'].iloc[-1]:.1f}",
                    f"{last_margin_adj:.1f}",
                ],
            })
            st.dataframe(compare, use_container_width=True, hide_index=True)
 
        st.divider()
 
        st.subheader("Eksport")
        ex1, ex2 = st.columns(2)
 
        with ex1:
            if st.button("Pobierz Raport PDF"):
                with st.spinner("Generowanie PDF..."):
                    pdf_raw = create_pdf(df, pred_adj, runway, fig_mc, st.session_state.ai_comment)
                    # ----------------------------------------------------------------------
                    b64 = base64.b64encode(pdf_raw).decode()
                    st.markdown(
                        f'<a href="data:application/octet-stream;base64,{b64}" '
                        f'download="raport.pdf">Zapisz PDF</a>',
                        unsafe_allow_html=True,
                    )
 
        with ex2:
            xlsx_bytes = to_excel(df)
            st.download_button(
                label="Pobierz Excel",
                data=xlsx_bytes,
                file_name="dane_finansowe.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
 
    except Exception as e:
        st.error(f"Wystapil blad: {e}")
        st.exception(e)
 
else:
    st.info("Wgraj plik CSV w panelu bocznym, aby odblokowac dashboard.")